# utils/excel_utils.py

import os
import time
import pandas as pd
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from datetime import datetime
from config import PRAKTIS_SEARCH_URL, PRAKTIKER_SEARCH_URL

def process_excel_and_split_files(input_file):
    """
    Reads the input Excel file (with three columns: Praktis Code, Praktiker Code, Buyer Code),
    fetches product data concurrently for unique product pairs,
    and returns:
      - product_data: a list of dictionaries for each unique product pair
      - buyer_mappings: a list of dictionaries mapping (Praktis Code, Praktiker Code) to Buyer Code
    """
    try:
        df = pd.read_excel(input_file, engine="odf")
        df_sorted = df.sort_values(by=df.columns[0])
        rows = df_sorted.values.tolist()
        buyer_mappings = []
        unique_pairs = {}
        for row in rows:
            praktis_code = str(row[0])
            praktiker_code = str(row[1])
            buyer_code = str(row[2])
            buyer_mappings.append({
                "Praktis Code": praktis_code,
                "Praktiker Code": praktiker_code,
                "Buyer Code": buyer_code
            })
            key = (praktis_code, praktiker_code)
            if key not in unique_pairs:
                unique_pairs[key] = {"Praktis Code": praktis_code, "Praktiker Code": praktiker_code}
        product_pairs = list(unique_pairs.values())
        results = []
        from scraping.scraping_functions import fetch_product_data_praktis, fetch_product_data_praktiker
        with ThreadPoolExecutor(max_workers=5) as executor:
            futures = {executor.submit(fetch_product_data_praktis, pair["Praktis Code"]): pair for pair in product_pairs}
            for future in as_completed(futures):
                pair = futures[future]
                praktis_data = future.result()
                praktiker_data = fetch_product_data_praktiker(pair["Praktiker Code"])
                result = {
                    "Praktis Code": str(pair["Praktis Code"]),
                    "Praktiker Code": str(pair["Praktiker Code"]),
                    "Praktis Name": str(praktis_data["name"]),
                    "Praktiker Name": str(praktiker_data["name"]),
                    "Praktis Regular Price": str(praktis_data["regular_price"]),
                    "Praktiker Regular Price": str(praktiker_data["regular_price"]),
                    "Praktis Promo Price": str(praktis_data["promo_price"]),
                    "Praktiker Promo Price": str(praktiker_data["promo_price"]),
                }
                results.append(result)
        results_sorted = sorted(results, key=lambda x: (x["Praktis Code"], x["Praktiker Code"]))
        return results_sorted, buyer_mappings
    except Exception as e:
        print(f"An error occurred while processing Excel: {e}")
        return [], []

def adjust_excel_formatting(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active
    for col in sheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
                cell.alignment = Alignment(wrap_text=True)
            except Exception as e:
                print(f"Error adjusting column {col_letter}: {e}")
        sheet.column_dimensions[col_letter].width = max_length + 2
    workbook.save(file_path)
    workbook.close()

def write_filtered_excel(file_path, filtered_data, buyer_info):
    """
    Writes the filtered product data to an Excel file.
    Adds a column "Buyer Info" containing buyer code, name, and email.
    Uses the same formatting as the original global Excel output (with hyperlinks).
    """
    buyer_str = f"{buyer_info.get('buyer_code','')} - {buyer_info.get('name','')} - {buyer_info.get('email','')}"
    for rec in filtered_data:
        rec["Buyer Info"] = buyer_str
    df = pd.DataFrame(filtered_data)
    with pd.ExcelWriter(file_path, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, sheet_name="Product Details")
        workbook = writer.book
        worksheet = writer.sheets["Product Details"]
        # Add hyperlinks on the product name columns.
        for row_num, row_data in enumerate(filtered_data, start=1):
            if row_data["Praktis Name"]:
                url = PRAKTIS_SEARCH_URL.format(row_data["Praktis Code"])
                worksheet.write_url(row_num, 2, url, string=row_data["Praktis Name"])
            if row_data["Praktiker Name"]:
                url = PRAKTIKER_SEARCH_URL.format(row_data["Praktiker Code"])
                worksheet.write_url(row_num, 3, url, string=row_data["Praktiker Name"])
        for col_num, col_name in enumerate(df.columns):
            max_length = max([len(str(val)) for val in df[col_name].fillna("")] + [len(col_name)])
            worksheet.set_column(col_num, col_num, max_length + 2, writer.book.add_format({'text_wrap': True}))
    print(f"Filtered Excel file written to {file_path}")

def format_email_body_table_html(filtered_changes, filtered_product_data):
    """
    Build an HTML table (columns: ID, Product name, My price, Their Price, Comp Change, Diff)
    using the filtered_changes dictionary and filtered_product_data.
    """
    rows_html = ""
    from utils.helpers import safe_float
    prod_dict = { (rec["Praktis Code"], rec["Praktiker Code"]) : rec for rec in filtered_product_data }
    for change in filtered_changes.get("price_changes", []):
        key = (change["code"], change["praktiker_code"])
        rec = prod_dict.get(key, {})
        my_price = safe_float(change.get("praktis_new_price", 0))
        their_price = safe_float(change.get("praktiker_new_price", 0))
        comp_change = safe_float(change.get("praktiker_new_price", 0)) - safe_float(change.get("praktiker_old_price", 0))
        diff = their_price - my_price
        name = rec.get("Praktis Name", "N/A")
        diff_str = f"<span style='color:red;'>-{abs(diff):.2f}</span>" if diff < 0 else (f"<span style='color:green;'>+{diff:.2f}</span>" if diff > 0 else f"{diff:.2f}")
        row_html = f"""
            <tr>
                <td style="border: 1px solid #ddd; padding: 8px;">{change['code']}</td>
                <td style="border: 1px solid #ddd; padding: 8px;">{name}</td>
                <td style="border: 1px solid #ddd; padding: 8px; text-align: right;">{my_price:.2f}</td>
                <td style="border: 1px solid #ddd; padding: 8px; text-align: right;">{their_price:.2f}</td>
                <td style="border: 1px solid #ddd; padding: 8px; text-align: right;">{comp_change:+.2f}</td>
                <td style="border: 1px solid #ddd; padding: 8px; text-align: right;">{diff_str}</td>
            </tr>
        """
        rows_html += row_html
    for update in filtered_changes.get("new_items", []):
        key = (update["Praktis Code"], update["Praktiker Code"])
        rec = prod_dict.get(key, {})
        my_price = safe_float(rec.get("Praktis Regular Price", 0))
        their_price = safe_float(rec.get("Praktiker Regular Price", 0))
        comp_change = 0.0
        diff = their_price - my_price
        name = rec.get("Praktis Name", "N/A")
        diff_str = f"<span style='color:red;'>-{abs(diff):.2f}</span>" if diff < 0 else (f"<span style='color:green;'>+{diff:.2f}</span>" if diff > 0 else f"{diff:.2f}")
        row_html = f"""
            <tr>
                <td style="border: 1px solid #ddd; padding: 8px;">{update['Praktis Code']}</td>
                <td style="border: 1px solid #ddd; padding: 8px;">{name}</td>
                <td style="border: 1px solid #ddd; padding: 8px; text-align: right;">{my_price:.2f}</td>
                <td style="border: 1px solid #ddd; padding: 8px; text-align: right;">{their_price:.2f}</td>
                <td style="border: 1px solid #ddd; padding: 8px; text-align: right;">{comp_change:+.2f}</td>
                <td style="border: 1px solid #ddd; padding: 8px; text-align: right;">{diff_str}</td>
            </tr>
        """
        rows_html += row_html
    table_html = f"""
    <html>
    <head>
      <style>
        table {{
          border-collapse: collapse;
          width: 100%;
          font-family: Arial, sans-serif;
        }}
        th, td {{
          border: 1px solid #ddd;
          padding: 8px;
          text-align: center;
        }}
        th {{
          background-color: #f2f2f2;
        }}
      </style>
    </head>
    <body>
      <h2>Price Comparison Report</h2>
      <table>
        <thead>
          <tr>
            <th>ID</th>
            <th>Product name</th>
            <th>My price</th>
            <th>Their Price</th>
            <th>Comp Change</th>
            <th>Diff</th>
          </tr>
        </thead>
        <tbody>
          {rows_html}
        </tbody>
      </table>
    </body>
    </html>
    """
    return table_html

def filter_product_data_by_buyer(product_data, buyer_code, buyer_mapping_dict):
    filtered = []
    for rec in product_data:
        key = (rec["Praktis Code"], rec["Praktiker Code"])
        if buyer_code in buyer_mapping_dict.get(key, []):
            filtered.append(rec)
    return filtered
